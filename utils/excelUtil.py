# coding:utf-8
from openpyxl import Workbook, load_workbook
import sys

if __name__ == '__main__':
    outwb = Workbook()
    wo = outwb.active

    careerSheet = outwb.create_sheet('career', 0)
    # careerSheet['A1']= datetime.datetime.now()
    careerSheet.cell(row=1, column=1).value = 'A'
    careerSheet.cell(row=1, column=2).value = 'B'
    careerSheet.cell(row=1, column=3).value = 'c'

    careerSheet.cell(row=2, column=2).value = 20
    careerSheet.cell(row=2, column=3).value = 30
    careerSheet.append([1, 2, 3])
    careerSheet.append(['This is A1', 'This is B1', 'This is C1'])
    careerSheet.append({'A': 'This is A1', 'C': 'This is C1'})
    careerSheet.append({1: 'This is A1', 3: 'This is C1'})

    outwb.save(r"test.xlsx")

    inwb = load_workbook("scname-test.xlsx")
    sheet = inwb.active
    scname = sheet["A6"].value
    print scname
    # print sys.path
