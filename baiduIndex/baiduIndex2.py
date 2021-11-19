# coding:utf-8
import datetime
import os
import random
import time
import urllib

import requests

from openpyxl import Workbook, load_workbook
from openpyxl.styles import *

import sys

import warnings

warnings.filterwarnings('ignore')


# import execjs


# 搜索指数数据解密
def decryption(keys, data):
    dec_dict = {}
    for j in range(len(keys) // 2):
        dec_dict[keys[j]] = keys[len(keys) // 2 + j]

    dec_data = ''
    for k in range(len(data)):
        dec_data += dec_dict[data[k]]
    return dec_data


if __name__ == "__main__":
    notIncluded = [];
    inwb = load_workbook("scname.xlsx")
    sheet = inwb.active
    inwbIndex = 2;
    ls = []
    scname = sheet["A" + str(inwbIndex)].value
    while (scname is not None):
        time.sleep(random.randint(3, 15))
        print scname + '=====================' + time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
        dataUrl = 'https://index.baidu.com/api/SearchApi/index?area=0&word=[[%7B%22name%22:%22' + scname + '%22,%22wordType%22:1%7D]]&startDate=2021-01-01&endDate=2021-10-31'
        keyUrl = 'https://index.baidu.com/Interface/ptbk?uniqid='
        header = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive',
            # 'Cookie': '你登陆之后的Cookie',
            'Cookie': 'BIDUPSID=A8A38EEC556F4B2B248B1FE20D7E271E; PSTM=1636334110; BAIDUID=A8A38EEC556F4B2B7BB0573EDF3CC507:FG=1; BDORZ=B490B5EBF6F3CD402E515D22BCDA1598; BDSFRCVID=OnAOJexroG0k-YnHih-R-efUWrpWxY5TDYrELPfiaimDVu-VJeC6EG0Pts1-dEu-EHtdogKK0mOTHv-F_2uxOjjg8UtVJeC6EG0Ptf8g0M5; H_BDCLCKID_SF=tR30WJbHMTrDHJTg5DTjhPrMXUTCWMT-MTryKKJoHxT_ObTdqqKVbtKA3p5jLbvkJGnRh4oNBUJtjJjYhfO45DuZyxomtfQxtNRJQKDE5p5hKq5S5-OobUPUDMc9LUvqHmcdot5yBbc8eIna5hjkbfJBQttjQn3hfIkj2CKLK-oj-D-RjT-23e; __yjs_duid=1_04e35b890a5b7a57aa2e97efbd12d64b1636334113926; Hm_lvt_d101ea4d2a5c67dab98251f0b5de24dc=1636334116; BDUSS=FWZ1FaMGJEaFVadklibmlFTXhJQU5SaUIyNlBDZXRpOGREREw3QjQxTXdCN0JoRVFBQUFBJCQAAAAAAAAAAAEAAAAXOl4xyPjArXN1bW1lcgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADB6iGEweohhbG; BDUSS_BFESS=FWZ1FaMGJEaFVadklibmlFTXhJQU5SaUIyNlBDZXRpOGREREw3QjQxTXdCN0JoRVFBQUFBJCQAAAAAAAAAAAEAAAAXOl4xyPjArXN1bW1lcgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADB6iGEweohhbG; CHKFORREG=adafa63f8cee635f34329a99de0c6a20; bdindexid=27s7t7ii4ubqi6p99r48gfuie3; Hm_lpvt_d101ea4d2a5c67dab98251f0b5de24dc=1636334406; ab_sr=1.0.1_NzRjOTE0ZTkxNWExNTllMzIzZmRlMWQxN2ZkOTcwZmE5NGZhZDFiNTBhZmFhMGE1Yjg4OWQxOWZlNzEyMGFiNTU5MDE0Mjc3NDBkZDhmYjMwZmIzODRlYTc5MmIyMjM0NGVjYjBlNTI2ZWFhYWM2MGM0ZjMzMmIzOWNlZTliMDkzNTZiMWE2MGUxYTA5NzZkZWI3NmQzOTczMzI3Zjg5Yg==; H_PS_PSSID=34833_34886_34990_34068_35053_34584_34504_34705_34916_34606_26350_34970; delPer=0; PSINO=1; BDSFRCVID_BFESS=OnAOJexroG0k-YnHih-R-efUWrpWxY5TDYrELPfiaimDVu-VJeC6EG0Pts1-dEu-EHtdogKK0mOTHv-F_2uxOjjg8UtVJeC6EG0Ptf8g0M5; H_BDCLCKID_SF_BFESS=tR30WJbHMTrDHJTg5DTjhPrMXUTCWMT-MTryKKJoHxT_ObTdqqKVbtKA3p5jLbvkJGnRh4oNBUJtjJjYhfO45DuZyxomtfQxtNRJQKDE5p5hKq5S5-OobUPUDMc9LUvqHmcdot5yBbc8eIna5hjkbfJBQttjQn3hfIkj2CKLK-oj-D-RjT-23e; RT="sl=0&ss=kvpz97mk&tt=0&bcn=https://fclog.baidu.com/log/weirwood?type=perf&z=1&dm=baidu.com&si=i2hkghhklgi&ul=1xs0v"',
            'Host': 'index.baidu.com',
            'Referer': 'https://index.baidu.com/v2/main/index.html',
            'sec-ch-ua': '" Not;A Brand";v="99", "Google Chrome";v="91", "Chromium";v="91"',
            'sec-ch-ua-mobile': '?0',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36'
        }
        # 设置请求超时时间为30秒
        resData = requests.get(dataUrl, timeout=30, headers=header)
        # 当前科学家没被百度指数收录，则记录进数组
        if (resData.json()['status'] != 0):
            notIncluded.append(scname)
            inwbIndex = inwbIndex + 1
            scname = sheet["A" + str(inwbIndex)].value
            continue
        uniqid = resData.json()['data']['uniqid']
        print("uniqid:{}".format(uniqid))
        keyData = requests.get(keyUrl + uniqid, timeout=30, headers=header)
        keyData.raise_for_status()
        keyData.encoding = resData.apparent_encoding

        # 开始对json数据进行解析
        startDate = resData.json()['data']['userIndexes'][0]['all']['startDate']
        print("startDate:{}".format(startDate))
        endDate = resData.json()['data']['userIndexes'][0]['all']['endDate']
        print("endDate:{}".format(endDate))
        source = (resData.json()['data']['userIndexes'][0]['all']['data'])  # 原加密数据
        print("原加密数据:{}".format(source))
        key = keyData.json()['data']  # 密钥
        print("密钥:{}".format(key))

        # Python的强大之处就在于，拥有很强大的第三方库，可以直接执行js代码，即对解密算法不熟悉，无法转换为Python代码时，直接执行js代码即可
        # js = execjs.compile('''
        #         function decryption(t, e){
        #             for(var a=t.split(""),i=e.split(""),n={},s=[],o=0;o<a.length/2;o++)
        #                 n[a[o]]=a[a.length/2+o]
        #             for(var r=0;r<e.length;r++)
        #                 s.push(n[i[r]])
        #             return s.join("")
        #         }
        # ''')
        # res = js.call('decryption', key, source)  # 调用此方式解密，需要打开上面的注解

        res = decryption(key, source)
        # print(type(res))
        resArr = res.split(",")

        dateStart = datetime.datetime.strptime(startDate, '%Y-%m-%d')
        dateEnd = datetime.datetime.strptime(endDate, '%Y-%m-%d')
        dataLs = []
        while dateStart <= dateEnd:
            dataLs.append(str(dateStart))
            dateStart += datetime.timedelta(days=1)
            # print(dateStart.strftime('%Y-%m-%d'))

        for i in range(len(dataLs)):
            print urllib.unquote(scname), dataLs[i], resArr[i]
            ls.append([urllib.unquote(scname), dataLs[i], resArr[i]])

        # wb = load_workbook('baiduIndex_scientist.xlsx');
        # ws1 = wb.create_sheet('科学家');
        inwbIndex = inwbIndex + 1
        scname = sheet["A" + str(inwbIndex)].value
        if os.path.exists('baiduIndex_scientist.xlsx'):
            outwb = load_workbook('baiduIndex_scientist.xlsx')
        else:
            outwb = Workbook()
        if 'result' in outwb.sheetnames:
            resultSheet =outwb.get_sheet_by_name('result')
        else:
            resultSheet = outwb.create_sheet('result', 0)
        for i in range(len(ls)):
            resultSheet.append(ls[i])
    notIncludedSheet = outwb.create_sheet('notIncluded', 1)
    notIncludedSheet.append(notIncluded)
    outwb.save("baiduIndex_scientist.xlsx")
